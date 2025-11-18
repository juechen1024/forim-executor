"""
Python adaptation of `ScheduleJobExecutor`.

Notes:
- This implementation keeps the original control flow and integration points
  (calls to inner clients) but simplifies some GIS-specific processing.
- Optional dependencies (install if you need full feature parity):
  - `openpyxl` for Excel: `pip install openpyxl`
  - `pyshp` (shapefile) and `shapely` for shapefile processing: `pip install pyshp shapely`
  - `rasterio` and `affine` for GeoTIFF processing: `pip install rasterio affine`
- The code uses existing Python client modules under `cn.gov.forestry.executor.client`.
- Some helpers from the Java code (FieldValueBuilder, ChineseUtils, CaseUtil, UuidCreator)
  are implemented with minimal behavior here to allow the executor to function.
"""
from dataclasses import dataclass
from typing import Any, Dict, List, Optional
import io
import uuid
import math
import tempfile
import zipfile
import os
import logging
import re
import unicodedata
from datetime import datetime

# Optional libraries - import lazily where used
from cn.gov.forestry.common.file.file_content import FileContent
from cn.gov.forestry.common.database.field_value import FieldValue
from cn.gov.forestry.common.database.crud.insert_batch_params import InsertBatchParams
from cn.gov.forestry.executor.client.general_inner_query_client import GeneralInnerQueryClient
from cn.gov.forestry.executor.client.assets_inner_resource_client import AssetsInnerResourceClient
from cn.gov.forestry.executor.client.database_inner_crud_client import DatabaseInnerCRUDClient
from cn.gov.forestry.executor.client.metadata_inner_query_client import MetadataInnerQueryClient
from cn.gov.forestry.executor.client.metadata_inner_opt_client import MetadataInnerOptClient
from cn.gov.forestry.executor.client.schedule_inner_job_client import ScheduleInnerJobClient
from cn.gov.forestry.common.domain.dto.schedule.schedule_job_dto import ScheduleJobDTO
from cn.gov.forestry.common.domain.dto.schedule.schedule_job_log_dto import ScheduleJobLogDTO
from cn.gov.forestry.common.domain.dto.metadata.metadata_table_dto import MetadataTableDTO
from cn.gov.forestry.common.domain.dto.metadata.metadata_field_dto import MetadataFieldDTO
from cn.gov.forestry.common.domain.dto.metadata.batch.metadata_field_batch_dto import MetadataFieldBatchDTO
from cn.gov.forestry.common.domain.dto.general.general_system_dto import GeneralSystemDTO
from cn.gov.forestry.common.database.database_info import DatabaseInfo

logger = logging.getLogger(__name__)


def _to_field_value(v: Any) -> FieldValue:
    # Minimal wrapper: store only the raw value.
    return FieldValue(dataTypeCode=None, value=v)


def _uuid_str() -> str:
    return str(uuid.uuid1())


class FieldValueBuilder:
    """Minimal re-implementation of Java FieldValueBuilder used by the executor.

    This provides helpers to create FieldValue objects for common use-cases
    and to convert a dict of plain values into a dict of FieldValue.
    """
    @staticmethod
    def string(v: Any) -> FieldValue:
        return FieldValue(dataTypeCode='string', value=str(v) if v is not None else None)

    @staticmethod
    def integer(v: Any) -> FieldValue:
        try:
            if v is None:
                val = None
            elif isinstance(v, int):
                val = v
            else:
                val = int(v)
        except Exception:
            val = None
        return FieldValue(dataTypeCode='integer', value=val)

    @staticmethod
    def double_value(v: Any) -> FieldValue:
        try:
            if v is None:
                val = None
            elif isinstance(v, float):
                val = v
            else:
                val = float(v)
        except Exception:
            val = None
        return FieldValue(dataTypeCode='double', value=val)

    @staticmethod
    def boolean_value(v: Any) -> FieldValue:
        if v is None:
            val = None
        elif isinstance(v, bool):
            val = v
        elif isinstance(v, str):
            val = v.strip().lower() in ('true', '1', 'yes', 'y')
        elif isinstance(v, (int, float)):
            val = bool(v)
        else:
            val = None
        return FieldValue(dataTypeCode='boolean', value=val)

    @staticmethod
    def date(v: Any) -> FieldValue:
        # try to parse common date formats, fallback to None or raw
        if v is None:
            return FieldValue(dataTypeCode='date', value=None)
        if isinstance(v, datetime):
            return FieldValue(dataTypeCode='date', value=v.isoformat())
        if isinstance(v, str):
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
                try:
                    dt = datetime.strptime(v, fmt)
                    return FieldValue(dataTypeCode='date', value=dt.isoformat())
                except Exception:
                    continue
            # fallback to iso parser
            try:
                dt = datetime.fromisoformat(v)
                return FieldValue(dataTypeCode='date', value=dt.isoformat())
            except Exception:
                return FieldValue(dataTypeCode='date', value=v)
        # other types: store raw
        return FieldValue(dataTypeCode='date', value=str(v))

    @staticmethod
    def object(v: Any) -> FieldValue:
        return FieldValue(dataTypeCode='object', value=v)

    @staticmethod
    def array(v: Any) -> FieldValue:
        return FieldValue(dataTypeCode='array', value=v)

    @staticmethod
    def create_field_value(data_type: Any, v: Any) -> FieldValue:
        """Create FieldValue by data_type which may be enum-like or a string code."""
        if data_type is None:
            return FieldValueBuilder.object(v)
        # accept string codes
        code = getattr(data_type, 'code', None) or (data_type if isinstance(data_type, str) else None)
        if code is None:
            return FieldValueBuilder.object(v)
        code = code.lower()
        if code in ('string', 'str'):
            return FieldValueBuilder.string(v)
        if code in ('integer', 'int'):
            return FieldValueBuilder.integer(v)
        if code in ('double', 'double_value', 'float'):
            return FieldValueBuilder.double_value(v)
        if code in ('boolean', 'bool'):
            return FieldValueBuilder.boolean_value(v)
        if code in ('date', 'datetime'):
            return FieldValueBuilder.date(v)
        if code in ('object', 'doc', 'document'):
            return FieldValueBuilder.object(v)
        if code in ('array', 'list'):
            return FieldValueBuilder.array(v)
        # fallback
        return FieldValueBuilder.object(v)

    @staticmethod
    def convert_object_map(m: Dict[str, Any]) -> Dict[str, FieldValue]:
        if not m:
            return {}
        return {k: FieldValueBuilder.object(v) for k, v in m.items()}

    @staticmethod
    def convert_with_metadata(m: Dict[str, Any], metadata_fields: List[MetadataFieldDTO]) -> Dict[str, FieldValue]:
        """Convert plain dict `m` into FieldValue dict guided by metadata_fields list."""
        if not m:
            return {}
        result: Dict[str, FieldValue] = {}
        meta_map = {f.fieldName: f for f in (metadata_fields or [])}
        for k, v in m.items():
            md = meta_map.get(k)
            dtype = md.fieldDataType if md is not None else None
            result[k] = FieldValueBuilder.create_field_value(dtype, v)
        return result


class ChineseUtils:
    """Utilities to convert Chinese keys to pinyin-like ASCII keys with uniqueness.

    This is a conservative implementation: if `pypinyin` is available it will be used,
    otherwise non-ascii characters are replaced by underscores. Duplicates get numeric suffixes.
    """
    @staticmethod
    def _simple_slug(s: str) -> str:
        # normalize and drop non-ascii, replace spaces with underscore
        s = unicodedata.normalize('NFKD', s)
        s = ''.join(ch for ch in s if ord(ch) < 128)
        s = re.sub(r'[^0-9a-zA-Z]+', '_', s).strip('_').lower()
        return s or 'field'

    @staticmethod
    def convert_keys_to_pinyin_with_unique_suffix(properties: Dict[str, Any]) -> Dict[str, Any]:
        try:
            from pypinyin import lazy_pinyin
            def to_pinyin(tok: str) -> str:
                return ''.join(lazy_pinyin(tok))
        except Exception:
            to_pinyin = None

        out: Dict[str, Any] = {}
        seen = {}
        for k, v in properties.items():
            if to_pinyin:
                new_k = to_pinyin(k)
                new_k = re.sub(r'[^0-9a-zA-Z]+', '_', new_k).lower()
                if not new_k:
                    new_k = ChineseUtils._simple_slug(k)
            else:
                new_k = ChineseUtils._simple_slug(k)

            base = new_k
            idx = seen.get(base, 0)
            if idx:
                new_k = f"{base}_{idx}"
            seen[base] = idx + 1
            out[new_k] = v
        return out


class CaseUtil:
    @staticmethod
    def _to_snake(name: str) -> str:
        s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
        s2 = re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1)
        s3 = s2.replace(' ', '_')
        s3 = re.sub(r'[^0-9a-zA-Z_]+', '_', s3)
        return s3.lower()

    @staticmethod
    def convert_keys_to_snake_case(m: Dict[str, FieldValue]) -> Dict[str, FieldValue]:
        out: Dict[str, FieldValue] = {}
        for k, v in m.items():
            out[CaseUtil._to_snake(k)] = v
        return out



def calculate_threshold(task_count: int) -> int:
    if task_count < 100:
        return 1
    elif task_count < 10_000:
        return 10
    elif task_count < 1_000_000:
        return 100
    elif task_count < 100_000_000:
        return 1000
    else:
        return 10_000


class ScheduleJobExecutor:
    def __init__(self,
                 general_client: GeneralInnerQueryClient,
                 assets_client: AssetsInnerResourceClient,
                 db_client: DatabaseInnerCRUDClient,
                 metadata_query_client: MetadataInnerQueryClient,
                 metadata_opt_client: MetadataInnerOptClient,
                 schedule_client: ScheduleInnerJobClient):
        self.general_client = general_client
        self.assets_client = assets_client
        self.db_client = db_client
        self.metadata_query_client = metadata_query_client
        self.metadata_opt_client = metadata_opt_client
        self.schedule_client = schedule_client

    def execute(self, job_dto: ScheduleJobDTO):
        logger.info(f"ScheduleJobExecutor received job {job_dto.id}")
        # update job status
        self.update_job_status(job_dto.systemId, job_dto.id, 'RUNNING')
        self.update_job_start_time(job_dto.systemId, job_dto.id)
        self.save_log(job_dto, 'INFO', 'execute:job received')
        try:
            job_type = job_dto.jobType
            if job_type and 'excel' in job_type.lower():
                self.execute_import_excel_job(job_dto)
            elif job_type and 'shapefile' in job_type.lower():
                self.execute_import_shapefile_job(job_dto)
            elif job_type and 'geotiff' in job_type.lower():
                self.execute_import_geotiff_job(job_dto)
            else:
                self.save_log(job_dto, 'ERROR', 'execute:job failed - unknown type')
                logger.info(f"Unknown job type for job {job_dto.id}")
                self.update_job_status(job_dto.systemId, job_dto.id, 'ERROR')
                return
            self.update_job_status(job_dto.systemId, job_dto.id, 'SUCCESS')
            self.save_log(job_dto, 'INFO', 'execute:job success')
        except Exception as e:
            logger.exception('ScheduleJobExecutor execute error')
            self.update_job_status(job_dto.systemId, job_dto.id, 'ERROR')
            self.update_job_result(job_dto.systemId, job_dto.id, {'error': str(e)})
            self.save_log(job_dto, 'ERROR', f'execute:job failed: {e}')
        finally:
            self.update_job_end_time(job_dto.systemId, job_dto.id)
            self.save_log(job_dto, 'INFO', 'execute:job stopped')

    def execute_import_excel_job(self, job_dto: ScheduleJobDTO):
        # Lazy import openpyxl
        try:
            from openpyxl import load_workbook
        except Exception:
            raise RuntimeError('openpyxl required for Excel import; install with pip install openpyxl')

        logger.info('execute_import_excel_job start')
        pre_offset = int(job_dto.jobTaskOffset or 0)
        params = job_dto.jobParams or {}
        system_id = params.get('systemId')
        table_id = params.get('tableId')

        general_system: GeneralSystemDTO = self.get_general_system_dto(system_id)
        database_info: DatabaseInfo = DatabaseInfo.from_dict(general_system.__dict__) if hasattr(DatabaseInfo, 'from_dict') else DatabaseInfo()
        metadata_table: MetadataTableDTO = self.get_metadata_table_dto(system_id, table_id)
        metadata_fields: List[MetadataFieldDTO] = self.get_metadata_field_list(system_id, table_id)

        excel_file = FileContent.from_dict(params.get('resourceFile')) if isinstance(params.get('resourceFile'), dict) else params.get('resourceFile')
        excel_file_content: FileContent = self.assets_client.get_resource_file(excel_file)

        wb = load_workbook(filename=io.BytesIO(excel_file_content.bytes), data_only=True, read_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows = list(sheet.rows)
        if not rows:
            return
        title_row = [cell.value for cell in rows[0]]
        task_count = max(0, len(rows) - 2)
        self.update_job_task_count(job_dto.systemId, job_dto.id, task_count)
        threshold = calculate_threshold(task_count)

        buffer: List[Dict[str, FieldValue]] = []
        offset = 2
        for idx in range(2, len(rows)):
            if offset > pre_offset:
                row = rows[idx]
                row_map: Dict[str, Any] = {}
                for j, cell in enumerate(row):
                    key = title_row[j] if j < len(title_row) else f'col_{j}'
                    val = cell.value
                    if val is None:
                        continue
                    row_map[str(key)] = val
                if row_map:
                    # convert to FieldValue map
                    insert_doc: Dict[str, FieldValue] = {}
                    for f in metadata_fields:
                        insert_doc[f.fieldName] = _to_field_value(row_map.get(f.fieldName))
                    # id
                    insert_doc['id'] = _to_field_value(_uuid_str())
                    buffer.append(insert_doc)
                if len(buffer) >= threshold:
                    self.insert_batch(database_info, metadata_table.tableEntityName, buffer)
                    buffer.clear()
                    task_offset = offset - 2 + 1
                    self.update_job_task_offset(job_dto.systemId, job_dto.id, task_offset)
                    self.save_log(job_dto, 'DEBUG', f'execute-excel:buffer batch sync offset:{task_offset}')
            offset += 1

        if buffer:
            self.insert_batch(database_info, metadata_table.tableEntityName, buffer)
            buffer.clear()
            task_offset = offset - 2
            self.update_job_task_offset(job_dto.systemId, job_dto.id, task_offset)
            self.save_log(job_dto, 'DEBUG', f'execute-excel:last buffer batch sync offset:{task_offset}')

    def execute_import_shapefile_job(self, job_dto: ScheduleJobDTO):
        # Lazy imports
        try:
            import shapefile  # pyshp
            from shapely.geometry import shape, mapping
        except Exception:
            raise RuntimeError('pyshp and shapely required for shapefile import; pip install pyshp shapely')

        logger.info('execute_import_shapefile_job start')
        pre_offset = int(job_dto.jobTaskOffset or 0)
        params = job_dto.jobParams or {}
        system_id = params.get('systemId')
        table_id = params.get('tableId')

        general_system: GeneralSystemDTO = self.get_general_system_dto(system_id)
        database_info: DatabaseInfo = DatabaseInfo.from_dict(general_system.__dict__) if hasattr(DatabaseInfo, 'from_dict') else DatabaseInfo()
        metadata_table: MetadataTableDTO = self.get_metadata_table_dto(system_id, table_id)

        shape_file = FileContent.from_dict(params.get('resourceFile')) if isinstance(params.get('resourceFile'), dict) else params.get('resourceFile')
        shape_file_content: FileContent = self.assets_client.get_resource_file(shape_file)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        try:
            tmp.write(shape_file_content.bytes)
            tmp.flush()
            tmp.close()
            extract_dir = tempfile.mkdtemp(prefix='shapefile-extract-')
            with zipfile.ZipFile(tmp.name, 'r') as zf:
                zf.extractall(extract_dir)

            # find .shp file
            shp_path = None
            for root, _, files in os.walk(extract_dir):
                for f in files:
                    if f.lower().endswith('.shp'):
                        shp_path = os.path.join(root, f)
                        break
                if shp_path:
                    break
            if not shp_path:
                raise RuntimeError('No .shp found in zip')

            reader = shapefile.Reader(shp_path)
            features = list(reader.iterShapeRecords())
            task_count = len(features)
            self.update_job_task_count(job_dto.systemId, job_dto.id, task_count)
            threshold = calculate_threshold(task_count)
            buffer: List[Dict[str, FieldValue]] = []
            offset = 0
            for rec in features:
                if offset > pre_offset:
                    geom = rec.shape.__geo_interface__
                    properties = rec.record.as_dict() if hasattr(rec.record, 'as_dict') else dict(zip([f[0] for f in reader.fields[1:]], rec.record))
                    # convert chinese keys -> pinyin/ascii with uniqueness
                    converted_properties = ChineseUtils.convert_keys_to_pinyin_with_unique_suffix(properties)
                    # convert property values to FieldValue objects
                    field_value_map = FieldValueBuilder.convert_object_map(converted_properties)
                    # convert keys to snake_case
                    properties_snake_case = CaseUtil.convert_keys_to_snake_case(field_value_map)

                    # append geometry and bbox as FieldValue objects
                    properties_snake_case[ 'geometry' ] = FieldValueBuilder.object(geom)
                    try:
                        from shapely.geometry import shape as _shape
                        bbox_val = _shape(geom).bounds
                        area_val = _shape(geom).area
                    except Exception:
                        bbox_val = None
                        area_val = None
                    properties_snake_case[ 'geometry_bbox' ] = FieldValueBuilder.object(bbox_val)
                    properties_snake_case[ 'geometry_area' ] = FieldValueBuilder.object(area_val)

                    # primary key
                    properties_snake_case['id'] = FieldValueBuilder.string(_uuid_str())

                    if offset == 0:
                        # auto create fields based on property keys
                        self.create_field_batch(system_id, table_id, properties_snake_case)

                    buffer.append(properties_snake_case)

                    if len(buffer) >= threshold:
                        self.insert_batch(database_info, metadata_table.tableEntityName, buffer)
                        buffer.clear()
                        task_offset = offset + 1
                        self.update_job_task_offset(job_dto.systemId, job_dto.id, task_offset)
                        self.save_log(job_dto, 'DEBUG', f'execute-shapefile:buffer batch sync offset:{task_offset}')
                offset += 1

            if buffer:
                self.insert_batch(database_info, metadata_table.tableEntityName, buffer)
                buffer.clear()
                task_offset = offset
                self.update_job_task_offset(job_dto.systemId, job_dto.id, task_offset)
                self.save_log(job_dto, 'DEBUG', f'execute-shapefile:last buffer batch sync offset:{task_offset}')
        finally:
            try:
                os.remove(tmp.name)
            except Exception:
                pass

    def execute_import_geotiff_job(self, job_dto: ScheduleJobDTO):
        # Lazy imports
        try:
            import rasterio
            from rasterio.windows import Window
            from affine import Affine
            from shapely.geometry import Polygon, mapping
        except Exception:
            raise RuntimeError('rasterio and affine required for geotiff import; pip install rasterio affine')

        logger.info('execute_import_geotiff_job start')
        pre_offset = int(job_dto.jobTaskOffset or 0)
        params = job_dto.jobParams or {}
        system_id = params.get('systemId')
        table_id = params.get('tableId')

        general_system: GeneralSystemDTO = self.get_general_system_dto(system_id)
        database_info: DatabaseInfo = DatabaseInfo.from_dict(general_system.__dict__) if hasattr(DatabaseInfo, 'from_dict') else DatabaseInfo()
        metadata_table: MetadataTableDTO = self.get_metadata_table_dto(system_id, table_id)

        tif_file = FileContent.from_dict(params.get('resourceFile')) if isinstance(params.get('resourceFile'), dict) else params.get('resourceFile')
        tif_file_content: FileContent = self.assets_client.get_resource_file(tif_file)

        with rasterio.MemoryFile(tif_file_content.bytes) as memfile:
            with memfile.open() as src:
                width = src.width
                height = src.height
                count = src.count
                transform = src.transform
                # compute task count
                task_count = width * height
                self.update_job_task_count(job_dto.systemId, job_dto.id, task_count)
                threshold = calculate_threshold(task_count)
                buffer: List[Dict[str, FieldValue]] = []
                offset = 0
                for row in range(height):
                    for col in range(width):
                        if offset > pre_offset:
                            values = [float(src.read(band + 1)[row, col]) for band in range(count)]
                            # skip nodata heuristics
                            if all(math.isnan(v) for v in values):
                                offset += 1
                                continue
                            # compute polygon for pixel
                            x, y = transform * (col - 0.5, row - 0.5)
                            pixel_width = transform.a
                            pixel_height = abs(transform.e)
                            coords = [
                                (x, y),
                                (x + pixel_width, y),
                                (x + pixel_width, y - pixel_height),
                                (x, y - pixel_height),
                                (x, y)
                            ]
                            polygon = Polygon(coords)
                            feature_props = {f'band_{i+1}': values[i] for i in range(count)}
                            # convert bands and create FieldValue map
                            converted_properties = {k: v for k, v in feature_props.items()}
                            field_value_map = FieldValueBuilder.convert_object_map(converted_properties)
                            # add geometry fields
                            field_value_map['geometry'] = FieldValueBuilder.object(mapping(polygon))
                            field_value_map['geometry_bbox'] = FieldValueBuilder.object(polygon.bounds)
                            field_value_map['geometry_area'] = FieldValueBuilder.object(polygon.area)
                            field_value_map['id'] = FieldValueBuilder.string(_uuid_str())

                            if offset == 0:
                                self.create_field_batch(system_id, table_id, field_value_map)

                            buffer.append(field_value_map)

                            if len(buffer) >= threshold:
                                self.insert_batch(database_info, metadata_table.tableEntityName, buffer)
                                buffer.clear()
                                task_offset = offset + 1
                                self.update_job_task_offset(job_dto.systemId, job_dto.id, task_offset)
                                self.save_log(job_dto, 'DEBUG', f'execute-geo-tiff:buffer batch sync offset:{task_offset}')
                        offset += 1
                if buffer:
                    self.insert_batch(database_info, metadata_table.tableEntityName, buffer)
                    buffer.clear()
                    task_offset = offset
                    self.update_job_task_offset(job_dto.systemId, job_dto.id, task_offset)
                    self.save_log(job_dto, 'DEBUG', f'execute-geo-tiff:last buffer batch sync offset:{task_offset}')

    # Helper methods to interact with inner services
    def update_job_status(self, system_id: str, job_id: str, status: str):
        dto = ScheduleJobDTO()
        dto.systemId = system_id
        dto.id = job_id
        dto.jobStatus = status
        try:
            self.schedule_client.update_asynchronous_job(dto)
        except Exception as e:
            logger.exception('update_job_status failed')

    def update_job_task_count(self, system_id: str, job_id: str, count: int):
        dto = ScheduleJobDTO()
        dto.systemId = system_id
        dto.id = job_id
        dto.jobTaskCount = count
        try:
            self.schedule_client.update_asynchronous_job(dto)
        except Exception:
            logger.exception('update_job_task_count failed')

    def update_job_task_offset(self, system_id: str, job_id: str, offset: int):
        dto = ScheduleJobDTO()
        dto.systemId = system_id
        dto.id = job_id
        dto.jobTaskOffset = offset
        try:
            self.schedule_client.update_asynchronous_job(dto)
        except Exception:
            logger.exception('update_job_task_offset failed')

    def update_job_start_time(self, system_id: str, job_id: str):
        dto = ScheduleJobDTO()
        dto.systemId = system_id
        dto.id = job_id
        dto.jobStartTime = datetime.now()
        try:
            self.schedule_client.update_asynchronous_job(dto)
        except Exception:
            logger.exception('update_job_start_time failed')

    def update_job_end_time(self, system_id: str, job_id: str):
        dto = ScheduleJobDTO()
        dto.systemId = system_id
        dto.id = job_id
        dto.jobEndTime = datetime.now()
        try:
            self.schedule_client.update_asynchronous_job(dto)
        except Exception:
            logger.exception('update_job_end_time failed')

    def update_job_result(self, system_id: str, job_id: str, result: Dict[str, Any]):
        dto = ScheduleJobDTO()
        dto.systemId = system_id
        dto.id = job_id
        dto.jobResult = result
        try:
            self.schedule_client.update_asynchronous_job(dto)
        except Exception:
            logger.exception('update_job_result failed')

    def get_general_system_dto(self, system_id: str) -> GeneralSystemDTO:
        query = GeneralSystemDTO()
        query.id = system_id
        return self.general_client.get_system_info(query)

    def get_metadata_table_dto(self, system_id: str, table_id: str) -> MetadataTableDTO:
        q = MetadataTableDTO()
        q.systemId = system_id
        q.id = table_id
        return self.metadata_query_client.get_metadata_table_info(q)

    def get_metadata_field_list(self, system_id: str, table_id: str) -> List[MetadataFieldDTO]:
        q = MetadataFieldDTO()
        q.systemId = system_id
        q.tableId = table_id
        return self.metadata_query_client.get_metadata_field_list_by_table(q)

    def insert_batch(self, database_info: DatabaseInfo, table_entity_name: str, properties_list: List[Dict[str, FieldValue]]) -> List[str]:
        params = InsertBatchParams()
        params.databaseInfo = database_info
        params.tableEntityName = table_entity_name
        params.propertiesList = properties_list
        try:
            return self.db_client.insert_batch(params)
        except Exception as e:
            logger.exception('insert_batch failed')
            return []

    def update_table_additional_properties(self, system_id: str, table_id: str, properties: Dict[str, Any]):
        dto = MetadataTableDTO()
        dto.systemId = system_id
        dto.id = table_id
        dto.additionalProperties = properties
        try:
            return self.metadata_opt_client.update_metadata_table_additional_properties(dto)
        except Exception:
            logger.exception('update_table_additional_properties failed')
            return None

    def create_field_batch(self, system_id: str, table_id: str, properties: Dict[str, FieldValue]) -> Optional[MetadataFieldBatchDTO]:
        dto = MetadataFieldBatchDTO()
        dto.systemId = system_id
        dto.tableId = table_id
        fields = []
        for key in properties.keys():
            f = MetadataFieldDTO()
            f.fieldName = key
            fields.append(f)
        dto.fields = fields
        try:
            return self.metadata_opt_client.create_field_batch(dto)
        except Exception:
            logger.exception('create_field_batch failed')
            return None

    def save_log(self, job_dto: ScheduleJobDTO, level: str, content: str):
        log = ScheduleJobLogDTO()
        log.systemId = job_dto.systemId
        log.jobId = job_dto.id
        log.jobType = job_dto.jobType
        log.jobExecutor = 'executor'
        log.jobLogLevel = 20 if level == 'INFO' else 40
        log.jobLogContent = content
        log.jobLogTime = datetime.now()
        try:
            self.schedule_client.create_schedule_job_log(log)
        except Exception:
            logger.exception('save_log failed')


__all__ = ['ScheduleJobExecutor']


def make_bbox_from_bounds(bounds) -> List[Optional[float]]:
    """Return [minx, miny, maxx, maxy] from a bounds tuple or sequence."""
    if not bounds:
        return [None, None, None, None]
    try:
        minx, miny, maxx, maxy = bounds
        return [float(minx), float(miny), float(maxx), float(maxy)]
    except Exception:
        return [None, None, None, None]


def shapefile_is_validated_dir(path: str) -> bool:
    has_shp = False
    has_dbf = False
    for root, _, files in os.walk(path):
        for f in files:
            if f.lower().endswith('.shp'):
                has_shp = True
            if f.lower().endswith('.dbf'):
                has_dbf = True
    return has_shp and has_dbf


def find_actual_extract_dir(extract_dir: str) -> str:
    # If there's a single directory inside extract_dir, return it, otherwise return extract_dir
    try:
        entries = [os.path.join(extract_dir, e) for e in os.listdir(extract_dir)]
        dirs = [p for p in entries if os.path.isdir(p)]
        if len(dirs) == 1:
            return dirs[0]
    except Exception:
        pass
    return extract_dir


def unzip_file(zip_path: str, output_dir: str):
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(output_dir)


def cleanup_path(path: str):
    try:
        if os.path.isfile(path):
            os.remove(path)
        elif os.path.isdir(path):
            for root, dirs, files in os.walk(path, topdown=False):
                for f in files:
                    try:
                        os.remove(os.path.join(root, f))
                    except Exception:
                        pass
                for d in dirs:
                    try:
                        os.rmdir(os.path.join(root, d))
                    except Exception:
                        pass
            try:
                os.rmdir(path)
            except Exception:
                pass
    except Exception as e:
        logger.exception('cleanup_path failed')


# Backwards-compatible camelCase wrappers to match the Java API names
def _make_feature_object_map(feature: Any) -> Dict[str, Any]:
    # try to produce a geojson-like dict with properties, geometry, bbox, area
    try:
        # if feature has __geo_interface__ (shapely/pygeo), use it
        if hasattr(feature, '__geo_interface__'):
            geo = feature.__geo_interface__
            props = {}
            # shapely geometry doesn't carry properties; caller should provide
            obj = {'type': 'Feature', 'geometry': geo, 'properties': props}
            # bbox
            try:
                from shapely.geometry import shape as _shape
                shapely_geom = _shape(geo)
                obj['properties']['geometry_bbox'] = make_bbox_from_bounds(shapely_geom.bounds)
                obj['properties']['geometry_area'] = shapely_geom.area
            except Exception:
                obj['properties']['geometry_bbox'] = None
                obj['properties']['geometry_area'] = None
            return obj
        # if it's a pyshp shapeRecord
        if hasattr(feature, 'shape') and hasattr(feature, 'record'):
            geom = feature.shape.__geo_interface__
            props = feature.record.as_dict() if hasattr(feature.record, 'as_dict') else dict(feature.record)
            obj = {'type': 'Feature', 'geometry': geom, 'properties': props}
            try:
                from shapely.geometry import shape as _shape
                shapely_geom = _shape(geom)
                obj['properties']['geometry_bbox'] = make_bbox_from_bounds(shapely_geom.bounds)
                obj['properties']['geometry_area'] = shapely_geom.area
            except Exception:
                obj['properties']['geometry_bbox'] = None
                obj['properties']['geometry_area'] = None
            return obj
    except Exception:
        logger.exception('make_feature_object_map failed')
    return {}


# Attach camelCase methods to ScheduleJobExecutor as simple delegators
def _attach_camel_case_aliases():
    SC = ScheduleJobExecutor
    # execution methods
    SC.executeImportExcelJob = SC.execute_import_excel_job
    SC.executeImportShapefileJob = SC.execute_import_shapefile_job
    SC.executeImportGeoTiffJob = SC.execute_import_geotiff_job
    # job update methods
    SC.updateJobStatus = SC.update_job_status
    SC.updateJobTaskCount = SC.update_job_task_count
    SC.updateJobTaskOffset = SC.update_job_task_offset
    SC.updateJobStartTime = SC.update_job_start_time
    SC.updateJobEndTime = SC.update_job_end_time
    SC.updateJobResult = SC.update_job_result
    # metadata/db helpers
    SC.getGeneralSystemDTO = SC.get_general_system_dto
    SC.getMetadataTableDTO = SC.get_metadata_table_dto
    SC.getMetadataFieldDTOList = SC.get_metadata_field_list
    SC.insertBatch = SC.insert_batch
    SC.updateTableAdditionalProperties = SC.update_table_additional_properties
    SC.createFieldBatch = SC.create_field_batch
    SC.saveLog = SC.save_log
    # file/shapefile helpers
    SC.makeFeatureObjectMap = staticmethod(_make_feature_object_map)
    SC.makeBbox = staticmethod(make_bbox_from_bounds)
    SC.shapefileIsValidated = staticmethod(shapefile_is_validated_dir)
    SC.unzip = staticmethod(unzip_file)
    SC.findActualExtractDir = staticmethod(find_actual_extract_dir)
    SC.cleanup = staticmethod(cleanup_path)


_attach_camel_case_aliases()
