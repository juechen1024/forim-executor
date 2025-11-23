import io
import json
import logging
import math
import os
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from cn.gov.forestry.common.database.database_info import DatabaseInfo
from cn.gov.forestry.common.database.field_value import FieldValue
from cn.gov.forestry.common.database.field_value_builder import FieldValueBuilder
from cn.gov.forestry.common.database.crud.insert_batch_params import InsertBatchParams
from cn.gov.forestry.common.domain.bo import (
	SystemScheduleJobStatusEnum,
	SystemScheduleJobTypeEnum,
	GeneralLogLevelEnum,
	SystemBuildInFieldEnum,
	SystemTableAdditionalPropertiesKeyEnum,
	SystemFieldDataTypeEnum,
)
from cn.gov.forestry.common.domain.dto.general.general_system_dto import GeneralSystemDTO
from cn.gov.forestry.common.domain.dto.metadata.metadata_table_dto import MetadataTableDTO
from cn.gov.forestry.common.domain.dto.metadata.metadata_field_dto import MetadataFieldDTO
from cn.gov.forestry.common.domain.dto.metadata.batch.metadata_field_batch_dto import MetadataFieldBatchDTO
from cn.gov.forestry.common.domain.dto.schedule.schedule_job_dto import ScheduleJobDTO
from cn.gov.forestry.common.domain.dto.schedule.schedule_job_log_dto import ScheduleJobLogDTO
from cn.gov.forestry.common.file.file_content import FileContent
from cn.gov.forestry.common.utils import ChineseUtils, CaseUtil
from cn.gov.forestry.executor.client.GeneralInnerQueryClient import GeneralInnerQueryClient
from cn.gov.forestry.executor.client.AssetsInnerResourceClient import AssetsInnerResourceClient
from cn.gov.forestry.executor.client.DatabaseInnerCRUDClient import DatabaseInnerCRUDClient
from cn.gov.forestry.executor.client.MetadataInnerQueryClient import MetadataInnerQueryClient
from cn.gov.forestry.executor.client.MetadataInnerOptClient import MetadataInnerOptClient
from cn.gov.forestry.executor.client.ScheduleInnerJobClient import ScheduleInnerJobClient

logger = logging.getLogger(__name__)


class ScheduleJobExecutor:
	def __init__(
		self,
		general_inner_query_client: GeneralInnerQueryClient,
		assets_inner_resource_client: AssetsInnerResourceClient,
		database_inner_crud_client: DatabaseInnerCRUDClient,
		metadata_inner_query_client: MetadataInnerQueryClient,
		metadata_inner_opt_client: MetadataInnerOptClient,
		schedule_inner_job_client: ScheduleInnerJobClient,
	):
		self.general_inner_query_client = general_inner_query_client
		self.assets_inner_resource_client = assets_inner_resource_client
		self.database_inner_crud_client = database_inner_crud_client
		self.metadata_inner_query_client = metadata_inner_query_client
		self.metadata_inner_opt_client = metadata_inner_opt_client
		self.schedule_inner_job_client = schedule_inner_job_client

	def execute(self, jobDTO: ScheduleJobDTO) -> None:
		logger.info("ScheduleJobExecutor-received-job-[%s]-execute...", jobDTO.id)
		self.updateJobStatus(jobDTO.systemId, jobDTO.id, SystemScheduleJobStatusEnum.RUNNING)
		self.updateJobStartTime(jobDTO.systemId, jobDTO.id)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute:job received")
		try:
			if SystemScheduleJobTypeEnum.isImportExcel(jobDTO.jobType):
				self.executeImportExcelJob(jobDTO)
			elif SystemScheduleJobTypeEnum.isImportShapefile(jobDTO.jobType):
				self.executeImportShapefileJob(jobDTO)
			elif SystemScheduleJobTypeEnum.isImportGeoTiff(jobDTO.jobType):
				self.executeImportGeoTiffJob(jobDTO)
			else:
				self.saveLog(jobDTO, GeneralLogLevelEnum.ERROR, "execute:job failed")
				logger.info("ScheduleJobExecutor-UNKNOW-job-[%s]-type", jobDTO.id)
			logger.info("ScheduleJobExecutor-execute-job-[%s]-SUCCESS", jobDTO.id)
			self.updateJobStatus(jobDTO.systemId, jobDTO.id, SystemScheduleJobStatusEnum.SUCCESS)
			self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute:job success")
		except Exception as exc:
			logger.error("ScheduleJobExecutor-execute-job-[%s]-ERROR", jobDTO.id, exc_info=exc)
			self.updateJobStatus(jobDTO.systemId, jobDTO.id, SystemScheduleJobStatusEnum.ERROR)
			self.updateJobResult(jobDTO.systemId, jobDTO.id, {"error": str(exc)})
			self.saveLog(jobDTO, GeneralLogLevelEnum.ERROR, f"execute:job failed:{exc}")
		finally:
			self.updateJobEndTime(jobDTO.systemId, jobDTO.id)
			self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute:job stoped")

	def executeImportExcelJob(self, jobDTO: ScheduleJobDTO) -> None:
		logger.info(
			"ScheduleJobExecutor-executeImportExcelJob-running-at-offset-[%s]...",
			jobDTO.jobTaskOffset,
		)
		pre_offset = self._safe_int(jobDTO.jobTaskOffset)
		params = jobDTO.jobParams or {}

		system_id = params.get("systemId")
		table_id = params.get("tableId")

		general_system = self.getGeneralSystemDTO(system_id)
		database_info = DatabaseInfo.from_system_info(general_system)
		metadata_table = self.getMetadataTableDTO(system_id, table_id)
		metadata_fields = self.getMetadataFieldDTOList(system_id, table_id)

		resource_value = params.get("resourceFile")
		excel_file = self._parseFileContent(resource_value)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-excel:read excel file source")
		excel_file_content = self.assets_inner_resource_client.getResourceFile(excel_file)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-excel:get excel file")

		try:
			from openpyxl import load_workbook
		except ImportError as exc:
			raise RuntimeError(
				"openpyxl is required for Excel import. Install it via 'pip install openpyxl'."
			) from exc

		workbook = load_workbook(filename=io.BytesIO(excel_file_content.bytes or b""), data_only=True)
		sheet = workbook[workbook.sheetnames[0]]

		excel_row_count = sheet.max_row or 0
		task_count = max(excel_row_count - 2, 0)
		self.updateJobTaskCount(jobDTO.systemId, jobDTO.id, task_count)
		self.saveLog(jobDTO, GeneralLogLevelEnum.DEBUG, f"execute-excel:task count:{task_count}")
		threshold = self.calculateThreshold(task_count)
		self.saveLog(jobDTO, GeneralLogLevelEnum.DEBUG, f"execute-excel:sync threshold:{threshold}")

		buffer: List[Dict[str, FieldValue]] = []
		title_cells = None
		offset = 1
		for row_index, row in enumerate(sheet.iter_rows(values_only=False)):
			if row_index == 0:
				title_cells = list(row)
				continue
			if row_index == 1:
				continue

			offset = row_index
			if offset <= pre_offset:
				continue

			row_map: Dict[str, Any] = {}
			for idx, cell in enumerate(row):
				if title_cells is None or idx >= len(title_cells):
					continue
				header = self.getString(title_cells[idx])
				if not header:
					continue
				value = cell.value
				if value is None:
					continue
				if isinstance(value, str) and value.strip() == "":
					continue
				row_map[header] = value

			if not row_map:
				continue

			insert_doc: Dict[str, FieldValue] = {}
			for metadata_field in metadata_fields:
				field_name = metadata_field.fieldName
				if not field_name:
					continue
				data_type = SystemFieldDataTypeEnum.getByCodeOrDefault(metadata_field.fieldDataType)
				insert_doc[field_name] = FieldValueBuilder.createFieldValue(
					data_type, row_map.get(field_name)
				)
			insert_doc[SystemBuildInFieldEnum.ID.fieldName] = FieldValueBuilder.generateUuidField()

			buffer.append(insert_doc)
			if len(buffer) >= threshold:
				self.insertBatch(database_info, metadata_table.tableEntityName, buffer)
				buffer.clear()
				task_offset = offset - 2 + 1
				self.updateJobTaskOffset(jobDTO.systemId, jobDTO.id, task_offset)
				self.saveLog(
					jobDTO,
					GeneralLogLevelEnum.DEBUG,
					f"execute-excel:buffer batch sync offset:{task_offset}",
				)

		if buffer:
			self.insertBatch(database_info, metadata_table.tableEntityName, buffer)
			buffer.clear()
			task_offset = offset - 2
			self.updateJobTaskOffset(jobDTO.systemId, jobDTO.id, task_offset)
			self.saveLog(
				jobDTO,
				GeneralLogLevelEnum.DEBUG,
				f"execute-excel:last buffer batch sync offset:{offset}",
			)

	def executeImportShapefileJob(self, jobDTO: ScheduleJobDTO) -> None:
		logger.info(
			"ScheduleJobExecutor-executeImportShapefileJob-running-at-offset-[%s]...",
			jobDTO.jobTaskOffset,
		)
		pre_offset = self._safe_int(jobDTO.jobTaskOffset)
		params = jobDTO.jobParams or {}

		system_id = params.get("systemId")
		table_id = params.get("tableId")

		general_system = self.getGeneralSystemDTO(system_id)
		database_info = DatabaseInfo.from_system_info(general_system)
		metadata_table = self.getMetadataTableDTO(system_id, table_id)

		resource_value = params.get("resourceFile")
		shape_file = self._parseFileContent(resource_value)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-shapefile:read shapefile source")
		shape_file_content = self.assets_inner_resource_client.getResourceFile(shape_file)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-shapefile:get shapefile")

		temp_zip_path: Optional[Path] = None
		extract_dir: Optional[Path] = None
		try:
			temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
			try:
				temp_file.write(shape_file_content.bytes or b"")
			finally:
				temp_file.close()
			temp_zip_path = Path(temp_file.name)
			self.saveLog(jobDTO, GeneralLogLevelEnum.DEBUG, "execute-shapefile:create temp zip file")

			extract_dir = Path(tempfile.mkdtemp(prefix="shapefile-extract-"))
			self.unzip(temp_zip_path, extract_dir)
			self.saveLog(jobDTO, GeneralLogLevelEnum.DEBUG, "execute-shapefile:unzip temp zip file")

			actual_extract_dir = self.findActualExtractDir(extract_dir)
			if not self.shapefileIsValidated(actual_extract_dir):
				return

			shp_path = self._locateShp(actual_extract_dir)
			if shp_path is None:
				raise RuntimeError("No .shp file found in archive")
			self.saveLog(jobDTO, GeneralLogLevelEnum.DEBUG, "execute-shapefile:find .shp file")

			try:
				import shapefile  # type: ignore
				from shapely.geometry import shape as shapely_shape, mapping  # type: ignore
			except ImportError as exc:
				raise RuntimeError(
					"pyshp and shapely are required for shapefile import. "
					"Install them via 'pip install pyshp shapely'."
				) from exc

			reader = shapefile.Reader(str(shp_path), encoding="utf-8")
			try:
				field_names = [field[0] for field in reader.fields[1:]]
				feature_records = list(reader.iterShapeRecords())
				feature_type_name = getattr(reader, "shapeTypeName", "")
				self.saveLog(
					jobDTO,
					GeneralLogLevelEnum.DEBUG,
					"execute-shapefile:read shapefile feature source",
				)
				self.saveLog(
					jobDTO,
					GeneralLogLevelEnum.DEBUG,
					f"execute-shapefile:read shapefile feature type{feature_type_name}",
				)

				min_x = math.inf
				min_y = math.inf
				max_x = -math.inf
				max_y = -math.inf
				for record in feature_records:
					geometry = shapely_shape(record.shape.__geo_interface__)
					if geometry.is_empty:
						continue
					bounds = geometry.bounds
					min_x = min(min_x, bounds[0])
					min_y = min(min_y, bounds[1])
					max_x = max(max_x, bounds[2])
					max_y = max(max_y, bounds[3])

				task_count = len(feature_records)
				self.updateJobTaskCount(jobDTO.systemId, jobDTO.id, task_count)
				self.saveLog(jobDTO, GeneralLogLevelEnum.DEBUG, f"execute-shapefile:task count:{task_count}")
				threshold = self.calculateThreshold(task_count)
				self.saveLog(
					jobDTO,
					GeneralLogLevelEnum.DEBUG,
					f"execute-shapefile:sync threshold:{threshold}",
				)

				crs_wkt, epsg_code = self._readPrj(shp_path)
				self.saveLog(
					jobDTO,
					GeneralLogLevelEnum.DEBUG,
					f"execute-shapefile:read shapefile feature crs wkt:{crs_wkt}",
				)
				if epsg_code is not None:
					self.saveLog(
						jobDTO,
						GeneralLogLevelEnum.DEBUG,
						f"execute-shapefile:read shapefile feature epsg code:{epsg_code}",
					)

				shape_metadata: Dict[str, Any] = {
					"featureCount": task_count,
					"crsWkt": crs_wkt,
				}
				if epsg_code is not None:
					shape_metadata["epsgCode"] = epsg_code
				if not math.isinf(min_x):
					shape_metadata[SystemBuildInFieldEnum.GEOMETRY_BBOX.fieldName] = self.makeBbox(
						min_x, min_y, max_x, max_y
					)

				self.updateTableAdditionalProperties(
					system_id,
					table_id,
					{
						SystemTableAdditionalPropertiesKeyEnum.SHAPE_FILE_METADATA_PROPERTIES.key: shape_metadata,
					},
				)
				self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-shapefile:update table metadata")

				buffer: List[Dict[str, FieldValue]] = []
				offset = 0
				for record in feature_records:
					if offset > pre_offset:
						properties_raw = (
							record.record.as_dict()
							if hasattr(record.record, "as_dict")
							else dict(zip(field_names, record.record))
						)
						converted_properties = ChineseUtils.convertKeysToPinyinWithUniqueSuffix(properties_raw)
						field_value_map = FieldValueBuilder.convertObjectMap(converted_properties)
						properties_snake_case = CaseUtil.convertKeysToSnakeCase(field_value_map)
						properties_snake_case[
							SystemBuildInFieldEnum.ID.fieldName
						] = FieldValueBuilder.generateUuidField()

						if offset == 0:
							field_batch = self.createFieldBatch(system_id, table_id, dict(properties_snake_case))
							if field_batch is not None:
								self.saveLog(
									jobDTO,
									GeneralLogLevelEnum.INFO,
									f"execute-shapefile:created field:{field_batch.createdCount}",
								)

						geometry = shapely_shape(record.shape.__geo_interface__)
						if not geometry.is_empty:
							properties_snake_case[
								SystemBuildInFieldEnum.GEOMETRY.fieldName
							] = FieldValueBuilder.object(mapping(geometry))
							properties_snake_case[
								SystemBuildInFieldEnum.GEOMETRY_BBOX.fieldName
							] = FieldValueBuilder.object(list(geometry.bounds))
							properties_snake_case[
								SystemBuildInFieldEnum.GEOMETRY_AREA.fieldName
							] = FieldValueBuilder.object(geometry.area)

						buffer.append(properties_snake_case)

						if len(buffer) >= threshold:
							self.insertBatch(database_info, metadata_table.tableEntityName, buffer)
							buffer.clear()
							task_offset = offset + 1
							self.updateJobTaskOffset(jobDTO.systemId, jobDTO.id, task_offset)
							self.saveLog(
								jobDTO,
								GeneralLogLevelEnum.DEBUG,
								f"execute-shapefile:buffer batch sync offset:{task_offset}",
							)
					offset += 1

				if buffer:
					self.insertBatch(database_info, metadata_table.tableEntityName, buffer)
					buffer.clear()
					task_offset = offset
					self.updateJobTaskOffset(jobDTO.systemId, jobDTO.id, task_offset)
					self.saveLog(
						jobDTO,
						GeneralLogLevelEnum.DEBUG,
						f"execute-shapefile:last buffer batch sync offset:{task_offset}",
					)
			finally:
				reader.close()
		except Exception as exc:
			logger.error("ExchangeShapefile-import-ERROR", exc_info=exc)
			self.saveLog(jobDTO, GeneralLogLevelEnum.ERROR, f"execute-shapefile:error{exc}")
			raise
		finally:
			self.cleanup(temp_zip_path, extract_dir)

	def executeImportGeoTiffJob(self, jobDTO: ScheduleJobDTO) -> None:
		logger.info(
			"ScheduleJobExecutor-executeImportGeoTiffJob-running-at-offset-[%s]...",
			jobDTO.jobTaskOffset,
		)
		pre_offset = self._safe_int(jobDTO.jobTaskOffset)
		params = jobDTO.jobParams or {}

		system_id = params.get("systemId")
		table_id = params.get("tableId")

		general_system = self.getGeneralSystemDTO(system_id)
		database_info = DatabaseInfo.from_system_info(general_system)
		metadata_table = self.getMetadataTableDTO(system_id, table_id)

		resource_value = params.get("resourceFile")
		geo_tiff_file = self._parseFileContent(resource_value)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-geo-tiff:get tiff resource")
		geo_tiff_file_content = self.assets_inner_resource_client.getResourceFile(geo_tiff_file)
		self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-geo-tiff:get tiff file")

		try:
			import rasterio  # type: ignore
			from shapely.geometry import Polygon, mapping  # type: ignore
		except ImportError as exc:
			raise RuntimeError(
				"rasterio and shapely are required for GeoTIFF import. "
				"Install them via 'pip install rasterio shapely'."
			) from exc

		try:
			with rasterio.MemoryFile(geo_tiff_file_content.bytes or b"") as memfile:
				with memfile.open() as dataset:
					width = dataset.width
					height = dataset.height
					band_count = dataset.count
					self.saveLog(
						jobDTO,
						GeneralLogLevelEnum.DEBUG,
						f"execute-geo-tiff:read tiff raster band:{band_count}",
					)

					crs = dataset.crs
					crs_wkt = crs.to_wkt() if crs else ""
					self.saveLog(
						jobDTO,
						GeneralLogLevelEnum.DEBUG,
						f"execute-geo-tiff:read tiff crs wkt:{crs_wkt}",
					)
					epsg_code = None
					if crs:
						try:
							epsg_code = crs.to_epsg()
						except Exception:
							epsg_code = None

					transform = dataset.transform
					pixel_width = transform.a
					pixel_height = abs(transform.e)
					rotation_x = transform.b
					rotation_y = transform.d
					bounds = dataset.bounds
					min_x, min_y, max_x, max_y = bounds.left, bounds.bottom, bounds.right, bounds.top

					tif_metadata: Dict[str, Any] = {
						"rasterWidth": width,
						"rasterHeight": height,
						"rasterNumBands": band_count,
						"crsWkt": crs_wkt,
						"pixelWidth": pixel_width,
						"pixelHeight": pixel_height,
						"rotationX": rotation_x,
						"rotationY": rotation_y,
						SystemBuildInFieldEnum.GEOMETRY_BBOX.fieldName: self.makeBbox(
							min_x, min_y, max_x, max_y
						),
						"rasterMinX": 0,
						"rasterMinY": 0,
						"rasterSampleModelTranslateX": 0,
						"rasterSampleModelTranslateY": 0,
						"rasterNumDataElements": band_count,
					}
					if epsg_code is not None:
						tif_metadata["epsgCode"] = epsg_code

					self.updateTableAdditionalProperties(
						system_id,
						table_id,
						{
							SystemTableAdditionalPropertiesKeyEnum.TIF_METADATA_PROPERTIES.key: tif_metadata,
						},
					)
					self.saveLog(jobDTO, GeneralLogLevelEnum.INFO, "execute-geo-tiff:update table metadata")

					task_count = height * width
					self.updateJobTaskCount(jobDTO.systemId, jobDTO.id, task_count)
					self.saveLog(
						jobDTO,
						GeneralLogLevelEnum.DEBUG,
						f"execute-geo-tiff:task count:{task_count}",
					)
					threshold = self.calculateThreshold(task_count)
					self.saveLog(
						jobDTO,
						GeneralLogLevelEnum.DEBUG,
						f"execute-geo-tiff:sync threshold:{threshold}",
					)

					data = dataset.read()
					nodata_values = dataset.nodatavals
					buffer: List[Dict[str, FieldValue]] = []
					offset = 0

					for row in range(height):
						for col in range(width):
							if offset > pre_offset:
								pixel_values = [
									float(data[band_index, row, col]) for band_index in range(band_count)
								]
								if self._isNoData(pixel_values, nodata_values):
									offset += 1
									continue

								properties = {
									f"band_{band_index + 1}": pixel_values[band_index]
									for band_index in range(band_count)
								}
								field_value_map = FieldValueBuilder.convertObjectMap(properties)
								properties_snake_case = CaseUtil.convertKeysToSnakeCase(field_value_map)
								properties_snake_case[
									SystemBuildInFieldEnum.ID.fieldName
								] = FieldValueBuilder.generateUuidField()

								if offset == 0:
									field_batch = self.createFieldBatch(
										system_id, table_id, dict(properties_snake_case)
									)
									if field_batch is not None:
										self.saveLog(
											jobDTO,
											GeneralLogLevelEnum.INFO,
											f"execute-geo-tiff:created field:{field_batch.createdCount}",
										)

								x_left, y_top = transform * (col, row)
								x_right = x_left + transform.a
								y_bottom = y_top + transform.e
								polygon = Polygon(
									[
										(x_left, y_top),
										(x_right, y_top),
										(x_right, y_bottom),
										(x_left, y_bottom),
										(x_left, y_top),
									]
								)
								properties_snake_case[
									SystemBuildInFieldEnum.GEOMETRY.fieldName
								] = FieldValueBuilder.object(mapping(polygon))
								properties_snake_case[
									SystemBuildInFieldEnum.GEOMETRY_BBOX.fieldName
								] = FieldValueBuilder.object(list(polygon.bounds))
								properties_snake_case[
									SystemBuildInFieldEnum.GEOMETRY_AREA.fieldName
								] = FieldValueBuilder.object(polygon.area)

								buffer.append(properties_snake_case)

								if len(buffer) >= threshold:
									self.insertBatch(database_info, metadata_table.tableEntityName, buffer)
									buffer.clear()
									task_offset = offset + 1
									self.updateJobTaskOffset(jobDTO.systemId, jobDTO.id, task_offset)
									self.saveLog(
										jobDTO,
										GeneralLogLevelEnum.DEBUG,
										f"execute-geo-tiff:buffer batch sync offset:{task_offset}",
									)
							offset += 1

					if buffer:
						self.insertBatch(database_info, metadata_table.tableEntityName, buffer)
						buffer.clear()
						task_offset = offset
						self.updateJobTaskOffset(jobDTO.systemId, jobDTO.id, task_offset)
						self.saveLog(
							jobDTO,
							GeneralLogLevelEnum.DEBUG,
							f"execute-geo-tiff:last buffer batch sync offset:{task_offset}",
						)
		except Exception as exc:
			logger.error("ExchangeGeoTiff-readGeoTiff-ERROR", exc_info=exc)
			self.saveLog(jobDTO, GeneralLogLevelEnum.ERROR, f"execute-geotiff:error:{exc}")
			raise

	def getString(self, cell: Any) -> str:
		value = getattr(cell, "value", None)
		if value is None:
			return ""
		return str(value)

	def shapefileIsValidated(self, extract_dir: Path) -> bool:
		has_shp = False
		has_dbf = False
		for entry in extract_dir.iterdir():
			if entry.is_file() and entry.suffix.lower() == ".shp":
				has_shp = True
			if entry.is_file() and entry.suffix.lower() == ".dbf":
				has_dbf = True
		if not has_shp or not has_dbf:
			logger.error("importShapefile-no-SHP-or-DBF")
			return False
		return True

	def unzip(self, zip_path: Path, output_dir: Path) -> None:
		with zipfile.ZipFile(zip_path, "r") as zip_file:
			for entry in zip_file.infolist():
				target_path = output_dir / entry.filename
				if entry.is_dir():
					target_path.mkdir(parents=True, exist_ok=True)
				else:
					target_path.parent.mkdir(parents=True, exist_ok=True)
					with zip_file.open(entry) as source, open(target_path, "wb") as target:
						target.write(source.read())

	def findActualExtractDir(self, extract_dir: Path) -> Path:
		entries = list(extract_dir.iterdir())
		if len(entries) == 1 and entries[0].is_dir():
			return entries[0]
		return extract_dir

	def cleanup(self, temp_zip: Optional[Path], extract_dir: Optional[Path]) -> None:
		if temp_zip and temp_zip.exists():
			try:
				temp_zip.unlink()
			except Exception:
				logger.error("ExchangeShapefile-cleanup-ERROR-[%s]", temp_zip, exc_info=True)

		if extract_dir and extract_dir.exists():
			try:
				self.deleteDirectoryRecursively(extract_dir)
			except Exception:
				logger.error("ExchangeShapefile-cleanup-ERROR-[%s]", extract_dir, exc_info=True)

	def deleteDirectoryRecursively(self, directory: Path) -> None:
		for root, dirs, files in os.walk(directory, topdown=False):
			for file_name in files:
				path = Path(root) / file_name
				try:
					path.unlink()
				except Exception:
					logger.error(
						"ExchangeShapefile-deleteDirectoryRecursively-ERROR-[%s]",
						path,
						exc_info=True,
					)
			for dir_name in dirs:
				path = Path(root) / dir_name
				try:
					path.rmdir()
				except Exception:
					logger.error(
						"ExchangeShapefile-deleteDirectoryRecursively-ERROR-[%s]",
						path,
						exc_info=True,
					)
		try:
			directory.rmdir()
		except Exception:
			logger.error(
				"ExchangeShapefile-deleteDirectoryRecursively-ERROR-[%s]",
				directory,
				exc_info=True,
			)

	def _locateShp(self, root: Path) -> Optional[Path]:
		for dir_path, _, files in os.walk(root):
			for file_name in files:
				if file_name.lower().endswith(".shp"):
					return Path(dir_path) / file_name
		return None

	def _readPrj(self, shp_path: Path) -> Tuple[str, Optional[int]]:
		prj_path = shp_path.with_suffix(".prj")
		if not prj_path.exists():
			return "", None
		text = prj_path.read_text(encoding="utf-8").strip()
		epsg_code: Optional[int] = None
		if text:
			try:
				import pyproj  # type: ignore

				epsg_code = pyproj.CRS.from_wkt(text).to_epsg()
			except ImportError:
				logger.debug("pyproj not installed; skip EPSG detection for shapefile")
			except Exception:
				epsg_code = None
		return text, epsg_code

	def makeBbox(self, min_x: float, min_y: float, max_x: float, max_y: float) -> List[float]:
		return [min_x, min_y, max_x, max_y]

	def updateJobStatus(
		self,
		system_id: Optional[str],
		job_id: Optional[str],
		job_status: SystemScheduleJobStatusEnum,
	) -> None:
		dto = ScheduleJobDTO(systemId=system_id, id=job_id, jobStatus=job_status.code)
		self.schedule_inner_job_client.updateAsynchronousJob(dto)

	def updateJobTaskCount(self, system_id: Optional[str], job_id: Optional[str], task_count: int) -> None:
		dto = ScheduleJobDTO(systemId=system_id, id=job_id, jobTaskCount=task_count)
		self.schedule_inner_job_client.updateAsynchronousJob(dto)

	def updateJobTaskOffset(self, system_id: Optional[str], job_id: Optional[str], task_offset: int) -> None:
		dto = ScheduleJobDTO(systemId=system_id, id=job_id, jobTaskOffset=task_offset)
		self.schedule_inner_job_client.updateAsynchronousJob(dto)

	def calculateThreshold(self, task_count: int) -> int:
		if task_count < 100:
			return 1
		if task_count < 10_000:
			return 10
		if task_count < 1_000_000:
			return 100
		if task_count < 100_000_000:
			return 1000
		return 10_000

	def updateJobStartTime(self, system_id: Optional[str], job_id: Optional[str]) -> None:
		dto = ScheduleJobDTO(systemId=system_id, id=job_id, jobStartTime=datetime.now())
		self.schedule_inner_job_client.updateAsynchronousJob(dto)

	def updateJobEndTime(self, system_id: Optional[str], job_id: Optional[str]) -> None:
		dto = ScheduleJobDTO(systemId=system_id, id=job_id, jobEndTime=datetime.now())
		self.schedule_inner_job_client.updateAsynchronousJob(dto)

	def updateJobResult(self, system_id: Optional[str], job_id: Optional[str], job_result: Dict[str, Any]) -> None:
		dto = ScheduleJobDTO(systemId=system_id, id=job_id, jobResult=job_result)
		self.schedule_inner_job_client.updateAsynchronousJob(dto)

	def getGeneralSystemDTO(self, system_id: Optional[str]) -> GeneralSystemDTO:
		dto = GeneralSystemDTO(id=system_id)
		return self.general_inner_query_client.getSystemInfo(dto)

	def getMetadataTableDTO(self, system_id: Optional[str], table_id: Optional[str]) -> MetadataTableDTO:
		dto = MetadataTableDTO(systemId=system_id, id=table_id)
		return self.metadata_inner_query_client.getMetadataTableInfo(dto)

	def getMetadataFieldDTOList(self, system_id: Optional[str], table_id: Optional[str]) -> List[MetadataFieldDTO]:
		dto = MetadataFieldDTO(systemId=system_id, tableId=table_id)
		return self.metadata_inner_query_client.getMetadataFieldListByTable(dto)

	def insertBatch(
		self,
		database_info: DatabaseInfo,
		table_entity_name: Optional[str],
		properties_list: List[Dict[str, FieldValue]],
	) -> List[str]:
		params = InsertBatchParams(
			databaseInfo=database_info,
			tableEntityName=table_entity_name,
			propertiesList=properties_list,
		)
		return self.database_inner_crud_client.insertBatch(params)

	def updateTableAdditionalProperties(
		self,
		system_id: Optional[str],
		table_id: Optional[str],
		properties: Dict[str, Any],
	) -> MetadataTableDTO:
		dto = MetadataTableDTO(systemId=system_id, id=table_id, additionalProperties=properties)
		return self.metadata_inner_opt_client.updateMetadataTableAdditionalProperties(dto)

	def createFieldBatch(
		self,
		system_id: Optional[str],
		table_id: Optional[str],
		properties: Dict[str, FieldValue],
	) -> Optional[MetadataFieldBatchDTO]:
		fields = [MetadataFieldDTO(fieldName=key) for key in properties.keys()]
		dto = MetadataFieldBatchDTO(systemId=system_id, tableId=table_id, fields=fields)
		return self.metadata_inner_opt_client.createFieldBatch(dto)

	def saveLog(
		self,
		jobDTO: ScheduleJobDTO,
		logLevel: GeneralLogLevelEnum,
		logContent: str,
	) -> None:
		job_log_dto = ScheduleJobLogDTO(
			systemId=jobDTO.systemId,
			jobId=jobDTO.id,
			jobType=jobDTO.jobType,
			jobExecutor="executor",
			jobLogLevel=logLevel.level,
			jobLogContent=logContent,
			jobLogTime=datetime.now(),
		)
		try:
			self.schedule_inner_job_client.createScheduleJobLog(job_log_dto)
		except Exception as exc:
			logger.error("sync-log-error", exc_info=exc)

	@staticmethod
	def _safe_int(value: Optional[Any], default: int = -1) -> int:
		try:
			if value is None:
				return default
			return int(value)
		except (TypeError, ValueError):
			return default

	def _parseFileContent(self, value: Any) -> FileContent:
		if isinstance(value, FileContent):
			return value
		if isinstance(value, dict):
			return FileContent.from_dict(value)
		if isinstance(value, str):
			try:
				parsed = json.loads(value)
			except json.JSONDecodeError as exc:
				raise ValueError("resourceFile must be JSON serializable") from exc
			return FileContent.from_dict(parsed)
		raise ValueError("resourceFile must be a dict, JSON string, or FileContent instance")

	@staticmethod
	def _isNoData(
		pixel_values: List[float],
		nodata_values: Optional[Tuple[Optional[float], ...]],
	) -> bool:
		if not pixel_values:
			return True
		all_nan = True
		for value in pixel_values:
			if isinstance(value, float):
				if math.isnan(value):
					continue
				all_nan = False
				break
			all_nan = False
			break
		if all_nan:
			return True
		if not nodata_values:
			return False
		for band_index, nodata in enumerate(nodata_values):
			if nodata is None:
				continue
			if band_index >= len(pixel_values):
				continue
			value = pixel_values[band_index]
			if isinstance(value, float) and isinstance(nodata, float):
				if math.isnan(value) and math.isnan(nodata):
					return True
			if value == nodata:
				return True
		return False


__all__ = ["ScheduleJobExecutor"]
